"""
Excel Exporter
Export payroll data to Excel format
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export payroll data to Excel"""

    @staticmethod
    def export_payroll_period(records, period_data, output_path):
        """
        Export a complete payroll period to Excel

        Args:
            records: List of payroll records
            period_data: Period information dict
            output_path: Path to save Excel file

        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Paie"

        # Header styling
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:S1')
        title_cell = ws['A1']
        title_cell.value = f"BULLETIN DE PAIE - {period_data.get('period_start', '')} au {period_data.get('period_end', '')}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 30

        # Headers
        headers = [
            "N° Mle", "Nom Complet", "Fonction", "Catégorie", "Statut",
            "Salaire Base", "Ind. Transport", "Alloc. Familiale", "Ind. Responsabilité",
            "Prime Risque", "Ind. Monture", "Heures Sup", "Salaire Brut",
            "INPS Salarié", "AMO Salarié", "Impôt", "Avances/Prêts",
            "Net à Payer", "Coût Total"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        row_num = 4
        totals = {
            'base_salary': 0, 'ind_transport': 0, 'family_allowance': 0,
            'responsibility_allowance': 0, 'risk_premium': 0, 'vehicle_allowance': 0,
            'overtime_pay': 0, 'gross_salary': 0, 'inps_employee': 0,
            'amo_employee': 0, 'income_tax': 0, 'advances_loans_deduction': 0,
            'net_to_pay': 0, 'total_payroll_cost': 0
        }

        for record in records:
            ws.cell(row=row_num, column=1, value=record.get('employee_id', ''))
            ws.cell(row=row_num, column=2, value=record.get('full_name', ''))
            ws.cell(row=row_num, column=3, value=record.get('position', ''))
            ws.cell(row=row_num, column=4, value=record.get('category', ''))
            ws.cell(row=row_num, column=5, value=record.get('status_code', ''))

            # Salary components
            col = 6
            for key in ['base_salary', 'ind_transport', 'family_allowance',
                       'responsibility_allowance', 'risk_premium', 'vehicle_allowance',
                       'overtime_pay', 'gross_salary', 'inps_employee', 'amo_employee',
                       'income_tax', 'advances_loans_deduction', 'net_to_pay']:
                value = record.get(key, 0)
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                cell.border = thin_border
                totals[key] += value
                col += 1

            # Total cost
            total_cost = record.get('total_payroll_cost', 0)
            cell = ws.cell(row=row_num, column=col, value=total_cost)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border
            totals['total_payroll_cost'] += total_cost

            # Alternate row colors
            if row_num % 2 == 0:
                for c in range(1, 20):
                    ws.cell(row=row_num, column=c).fill = PatternFill(
                        start_color="ECF0F1", end_color="ECF0F1", fill_type="solid"
                    )

            row_num += 1

        # Totals row
        total_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        total_font = Font(color="FFFFFF", bold=True, size=11)

        ws.cell(row=row_num, column=1, value="TOTAUX").font = total_font
        ws.cell(row=row_num, column=1).fill = total_fill
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = total_fill
            ws.cell(row=row_num, column=col).border = thin_border

        col = 6
        for key in ['base_salary', 'ind_transport', 'family_allowance',
                   'responsibility_allowance', 'risk_premium', 'vehicle_allowance',
                   'overtime_pay', 'gross_salary', 'inps_employee', 'amo_employee',
                   'income_tax', 'advances_loans_deduction', 'net_to_pay', 'total_payroll_cost']:
            cell = ws.cell(row=row_num, column=col, value=totals[key])
            cell.fill = total_fill
            cell.font = total_font
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border
            col += 1

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        for col in range(6, 20):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Freeze panes
        ws.freeze_panes = 'A4'

        # Save
        wb.save(output_path)
        return output_path

    @staticmethod
    def export_bank_transfers(records, period_data, output_path):
        """
        Export bank transfer list to Excel

        Args:
            records: List of payroll records
            period_data: Period information dict
            output_path: Path to save Excel file

        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Virements"

        # Header styling
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"LISTE DES VIREMENTS BANCAIRES - {period_data.get('period_start', '')}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 25

        # Headers
        headers = ["N° Mle", "Nom Complet", "Banque", "N° Compte", "Net à Payer", "Date Virement", "Référence"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        row_num = 4
        total_amount = 0
        virement_date = datetime.now().strftime("%d/%m/%Y")

        for record in records:
            ws.cell(row=row_num, column=1, value=record.get('employee_id', ''))
            ws.cell(row=row_num, column=2, value=record.get('full_name', ''))
            ws.cell(row=row_num, column=3, value=record.get('bank_name', 'N/A'))
            ws.cell(row=row_num, column=4, value=record.get('account_number', 'N/A'))

            net = record.get('net_to_pay', 0)
            ws.cell(row=row_num, column=5, value=net).number_format = '#,##0'
            ws.cell(row=row_num, column=6, value=virement_date)
            ws.cell(row=row_num, column=7, value=f"VIR-{record.get('employee_id', '')}-{period_data.get('period_start', '')}")

            total_amount += net
            row_num += 1

        # Total
        total_row = row_num + 1
        ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
        cell = ws.cell(row=total_row, column=5, value=total_amount)
        cell.font = Font(bold=True, size=12)
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30

        wb.save(output_path)
        return output_path
